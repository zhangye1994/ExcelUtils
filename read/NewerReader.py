import openpyxl
import openpyxl.workbook as xlsx_work_book
import openpyxl.worksheet as xlsx_sheet

from read.Reader import Reader
from utils import Printer


class NewerReader(Reader):
    wb: xlsx_work_book = None
    sheets = {}

    def __init__(self, file_path):
        """
        初始化Reader，并打开指定工作薄

        :param file_path: sheet路径
        """
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)

    def get_sheet_names(self):
        """
        获取当前workbook所有的sheet名称

        :return: sheet名称集合
        """
        sheet_names = self.wb.get_sheet_names()
        return sheet_names

    def open_sheet(self, sheet_name):
        """
        打开指定名称Sheet

        :param sheet_name: sheet名称
        :return: 返回sheet
        """
        sheet: xlsx_sheet = self.wb[sheet_name]
        self.sheets[sheet_name] = sheet
        return sheet

    def get_rows(self, sheet_name):
        """
        获取excel表格行数

        :param sheet_name: sheet名称
        :return: 行数
        """
        sheet: xlsx_sheet = self.open_sheet(sheet_name)
        return sheet.max_row

    def get_columns(self, sheet_name: str):
        """
        获取excel表格列数

        :param sheet_name: sheet名称
        :return: 列数
        """
        sheet: xlsx_sheet = self.open_sheet(sheet_name)
        return sheet.max_column

    def get_columns_by_name(self, sheet_name: str, column_names, row_line=0):
        """
        获取指定行对应列值所在列号

        :param row_line: 从指定行读取对应的列值，默认首行
        :param sheet_name: sheet名称
        :param column_names: 列名
        :return: 以字典格式返回对应字段所在的列
        """
        columns = self.get_columns(sheet_name)
        column_index = {}
        for index in range(0, columns):
            column_value: str = self.get_single_value(sheet_name, row_line, index)
            column_index[column_value] = index
        column_nums = []
        for name in column_names:
            column_nums.append(column_index.get(name, -1))
        return column_nums

    def get_single_value(self, sheet_name: str, row, col) -> str:
        """
        根据行列号获取对应的值

        :param sheet_name: sheet名称
        :param row: 行号
        :param col: 列号
        :return: 对应的值，为空时返回空串，避免类型异常
        """
        if row < 0 or col < 0:
            Printer.print_warn("Row or column values must be at least 0")
            return ""
        row += 1
        col += 1
        sheet: xlsx_sheet = self.open_sheet(sheet_name)
        value = sheet.cell(row, col).value
        if value is None:
            return ""
        return str(value)

    def get_row_values(self, sheet_name: str):
        """
        按行读取Excel所有数据

        :param sheet_name: sheet名称
        :return: 二维数组格式Excel数据
        """
        sheet: xlsx_sheet = self.open_sheet(sheet_name)
        row_result = []
        row_values = sheet.iter_rows(values_only=True)
        for row_item in row_values:
            row_result.append(row_item)
        return row_result

    def get_column_values(self, sheet_name: str, column_index=None):
        """
        按列读取Excel所有数据

        :param sheet_name: sheet名称
        :param column_index: 对应的列名索引数组
        :return: 二维数组格式Excel数据
        """
        if column_index is None:
            Printer.print_warn("No column index specified")
            return
        sheet: xlsx_sheet = self.open_sheet(sheet_name)
        column_result = []
        row_values = sheet.iter_rows(values_only=True)
        for row_item in row_values:
            values = []
            for index in column_index:
                values.append(row_item[index])
            column_result.append(values)
        return column_result

    def get_all_value(self, sheet_name: str):
        """
        按行列逐个获取excel所有数据

        :param sheet_name: sheet名称
        :return: 一维数组格式Excel数据
        """
        sheet: xlsx_sheet = self.open_sheet(sheet_name)
        row_result = []
        row_values = sheet.iter_rows(values_only=True)
        for row_item in row_values:
            for value in row_item:
                row_result.append(value)
        return row_result

    def close(self):
        """
        关闭工作簿
        封装原始代码，闭环整个Reader

        :return: NA
        """
        self.wb.close()
